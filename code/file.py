from tkinter import filedialog, messagebox

import openpyxl
from openpyxl import Workbook


def openfile(data):
    filename = filedialog.askopenfilename(filetypes=[('xlsx', '*.xlsx')], initialdir='C:\\')
    # print(name)
    if filename != "":   #打开文件
        wb = openpyxl.load_workbook(filename)
        sheet = wb.worksheets[0]
        for i in range(sheet.max_row):
            stu = []
            for t in range(sheet.max_column):
                stu.append(sheet.cell(i + 1, t + 1).value)
            data[i] = stu
        return sheet.max_row
    else:
        return 0

def savefile(data):
    try:
        if  data:  # 判断是否爬取到数据，是否需要保存excel文件
            # biaoti = [['序号', '企业名称', '生产经营场所地址', '行业类别', '所在地区', '发证机关', '许可证编号', '办结日期', '有效期限', 'COD年排放量', '氨氮年排放量', '二氧化硫年排放量', '氮氧化物年排放量']]
            wb = Workbook()
            sheet = wb.worksheets[0]
            filename = filedialog.asksaveasfilename(filetypes=[('xlsx', '*.xlsx')], initialdir='C:\\')
            filename = filename + '.xlsx'
            for row_id in data.keys():
                for items in range(len(data[row_id])):
                    sheet.cell(row_id + 1, items + 1).value = data[row_id][items]
            if filename != '.xlsx':
                wb.save(filename)
                messagebox.showinfo("提示", "保存完毕~！！！")
        else:
            messagebox.showinfo("提示", '没有数据,不必保存')
    except:
        messagebox.showinfo("提示", '保存文件错误，请重试~！！')
